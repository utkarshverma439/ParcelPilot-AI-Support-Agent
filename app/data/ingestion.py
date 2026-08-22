import os
import json
import uuid
import pymupdf
import openpyxl
from sqlalchemy.orm import Session
from app.data.models import Account, Order, Ticket, Base
from app.vector.store import VectorStore
from app.services.reliability import DOCUMENT_METADATA

RAW_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw")


def _read_readme_timestamp(wb) -> str | None:
    for name in wb.sheetnames:
        if name.lower() in ("readme", "read me", "info", "metadata"):
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and "snapshot" in cell.lower():
                        idx = row.index(cell)
                        if idx + 1 < len(row) and row[idx + 1]:
                            return str(row[idx + 1])
                    if cell and isinstance(cell, str) and "timestamp" in cell.lower():
                        idx = row.index(cell)
                        if idx + 1 < len(row) and row[idx + 1]:
                            return str(row[idx + 1])
    return None


def _find_sheet(wb, preferred_names: list[str]) -> str | None:
    for name in preferred_names:
        if name in wb.sheetnames:
            return name
    for sheet_name in wb.sheetnames:
        lower = sheet_name.lower()
        for pref in preferred_names:
            if pref.lower() in lower:
                return sheet_name
    return None


def ingest_excel(db: Session):
    path = os.path.join(RAW_DIR, "ParcelPilot_Assessment_Data.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)

    snapshot_time = _read_readme_timestamp(wb)
    if snapshot_time:
        print(f"  README snapshot timestamp: {snapshot_time}")

    accounts_sheet_name = _find_sheet(wb, ["accounts", "account"])
    if accounts_sheet_name:
        accounts_sheet = wb[accounts_sheet_name]
        headers = [cell.value for cell in accounts_sheet[1]]
        for row in accounts_sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            account = Account(
                account_id=str(data["account_id"]),
                account_name=str(data["account_name"]),
                plan=str(data["plan"]),
                status=str(data["status"]),
                csm=str(data["csm"]) if data.get("csm") else None,
                contract_file=str(data["contract_file"]) if data.get("contract_file") else None,
                premium_support=bool(data.get("premium_support", False)),
                notes=str(data["notes"]) if data.get("notes") else None,
            )
            db.merge(account)
        print(f"  Ingested accounts from sheet: {accounts_sheet_name}")

    orders_sheet_name = _find_sheet(wb, ["orders", "order"])
    if orders_sheet_name:
        orders_sheet = wb[orders_sheet_name]
        headers = [cell.value for cell in orders_sheet[1]]
        for row in orders_sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            order = Order(
                order_id=str(data["order_id"]),
                account_id=str(data["account_id"]),
                carrier=str(data["carrier"]),
                status=str(data["status"]),
                booked_at=str(data["booked_at"]) if data.get("booked_at") else None,
                pickup_window_start=str(data["pickup_window_start"]) if data.get("pickup_window_start") else None,
                pickup_window_end=str(data["pickup_window_end"]) if data.get("pickup_window_end") else None,
                pickup_actual_at=str(data["pickup_actual_at"]) if data.get("pickup_actual_at") else None,
                shipment_fee_inr=float(data["shipment_fee_inr"]) if data.get("shipment_fee_inr") else 0.0,
                carrier_fault=bool(data.get("carrier_fault", False)),
                customer_fault=bool(data.get("customer_fault", False)),
                cancellation_requested_at=str(data["cancellation_requested_at"]) if data.get("cancellation_requested_at") else None,
                notes=str(data["notes"]) if data.get("notes") else None,
            )
            db.merge(order)
        print(f"  Ingested orders from sheet: {orders_sheet_name}")

    tickets_sheet_name = _find_sheet(wb, ["tickets", "ticket"])
    if tickets_sheet_name:
        tickets_sheet = wb[tickets_sheet_name]
        headers = [cell.value for cell in tickets_sheet[1]]
        for row in tickets_sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            ticket = Ticket(
                ticket_id=str(data["ticket_id"]),
                account_id=str(data["account_id"]),
                created_at=str(data["created_at"]) if data.get("created_at") else None,
                status=str(data["status"]),
                subject=str(data["subject"]) if data.get("subject") else None,
                description=str(data["description"]) if data.get("description") else None,
                channel=str(data["channel"]) if data.get("channel") else None,
                assigned_to=str(data["assigned_to"]) if data.get("assigned_to") else None,
                last_customer_message_at=str(data["last_customer_message_at"]) if data.get("last_customer_message_at") else None,
                historical_resolution=str(data["historical_resolution"]) if data.get("historical_resolution") else None,
            )
            db.merge(ticket)
        print(f"  Ingested tickets from sheet: {tickets_sheet_name}")

    db.commit()
    wb.close()


def ingest_pdfs(vector_store: VectorStore):
    pdf_files = [
        "01_Support_Policy_v3_CURRENT.pdf",
        "02_Support_Policy_v2_DEPRECATED.pdf",
        "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
        "04_Product_Operations_Guide_and_Known_Issues.pdf",
        "05_Northstar_Logistics_Enterprise_Agreement.pdf",
        "06_LumenWorks_Service_Agreement.pdf",
    ]

    for pdf_file in pdf_files:
        path = os.path.join(RAW_DIR, pdf_file)
        if not os.path.exists(path):
            print(f"  Skipping {pdf_file} - not found")
            continue

        meta = DOCUMENT_METADATA.get(pdf_file, {})
        doc = pymupdf.open(path)
        chunks = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            if not text.strip():
                continue

            paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
            if not paragraphs:
                paragraphs = [text.strip()]

            for i, paragraph in enumerate(paragraphs):
                if len(paragraph) < 10:
                    continue
                chunk_id = f"{pdf_file}:p{page_num + 1}:c{i}"
                chunks.append({
                    "id": chunk_id,
                    "text": paragraph,
                    "metadata": {
                        "document_name": meta.get("document_name", pdf_file),
                        "document_type": meta.get("document_type", "unknown"),
                        "version": meta.get("version", "unknown"),
                        "status": meta.get("status", "unknown"),
                        "effective_date": meta.get("effective_date", "unknown"),
                        "customer_account_id": meta.get("customer_account_id"),
                        "source_priority": meta.get("source_priority", 50),
                        "page_number": page_num + 1,
                        "section": meta.get("section", "general"),
                        "source_file": pdf_file,
                    },
                })

        doc.close()

        if chunks:
            vector_store.add_documents(chunks)
            print(f"  Ingested {pdf_file}: {len(chunks)} chunks")


def run_ingestion(db: Session, vector_store: VectorStore):
    print("Ingesting Excel data...")
    ingest_excel(db)
    print("Excel ingestion complete.")

    print("Ingesting PDF documents...")
    ingest_pdfs(vector_store)
    print("PDF ingestion complete.")
